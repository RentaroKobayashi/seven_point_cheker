"""AB分類スクリプト（独立実行）

正誤判定結果Excelを入力として、カタログ値の有無で項目をA/Bに振り分け、
Aシート（マスター照合結果）とBシート（カタログ外の記載内容 + 差分コメント）を
含むExcelを出力する。

処理フロー:
  1. 正誤判定Excelを読み込み（縦並びレイアウト）
  2. 各項目をカタログ値の有無でA/Bに振り分け（ルールベース、AI不要）
  3. A: 正誤判定をそのまま使用
  4. B: 各資料の記載内容を記録 + 2資料以上あれば一致/不一致判定
  5. B項目にAI差分コメント生成
  6. Excel出力（Aシート + Bシート）

使い方:
  uv run python ab_classifier.py <正誤判定Excelパス>
"""

import io
import json
import logging
import os
import sys
from datetime import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from dotenv import load_dotenv
from google import genai
from tenacity import retry, stop_after_attempt, wait_exponential

load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

# Gemini モデル
_MODEL = "gemini-3-flash-preview"

# 技術文書の表示名 → 内部キー のマッピング（Excel読み込み用）
_DISPLAY_TO_KEY = {
    "図面(組立図)": "図面",
    "承認図": "承認図",
    "器具銘板(現物)": "器具銘板",
    "取説": "取扱説明書",
    "外装表示": "外装ラベル",
}

# 内部キー → 表示名
_KEY_TO_DISPLAY = {v: k for k, v in _DISPLAY_TO_KEY.items()}

# 技術文書の表示順序
_DOC_TYPE_ORDER = ["図面", "承認図", "器具銘板", "取扱説明書", "外装ラベル"]


# ============================================================
# 1. Excel読み込み
# ============================================================

def parse_comparison_excel(excel_path: str) -> dict:
    """正誤判定Excelを読み込み、構造化データに変換する。

    Returns:
        {
            "field_names": ["品番", "電圧", ...],
            "catalog": {"品番": "XXX", "電圧": "100V", ...},
            "doc_types": ["図面", "承認図", ...],
            "data": {
                "図面": {
                    "品番": {"extracted": "...", "correct": "...", "status": "○/×/-"},
                    ...
                },
                ...
            }
        }
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)
    ws = wb["比較結果"]

    # ヘッダー行1を読み取り、技術文書の開始列を特定
    row1 = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # 技術文書の開始列: col 2 はカタログ、col 3以降が技術文書（3列セット）
    doc_columns: list[tuple[str, int]] = []  # (内部キー, 開始列index)
    for col_idx, val in enumerate(row1):
        if val and val in _DISPLAY_TO_KEY:
            doc_columns.append((_DISPLAY_TO_KEY[val], col_idx))

    # データ行を読み取り（3行目から）
    field_names: list[str] = []
    catalog: dict[str, str] = {}
    data: dict[str, dict[str, dict]] = {key: {} for key, _ in doc_columns}

    for row in ws.iter_rows(min_row=3, values_only=True):
        row_vals = list(row)
        if not row_vals or not row_vals[0]:
            continue

        field_name = str(row_vals[0]).strip()
        field_names.append(field_name)

        # カタログ値（列2 = index 1）
        cat_val = str(row_vals[1]).strip() if row_vals[1] is not None else "-"
        if cat_val in ("None", ""):
            cat_val = "-"
        catalog[field_name] = cat_val

        # 各技術文書: 抽出(+0), 正解(+1), 判定(+2)
        for doc_key, start_col in doc_columns:
            extracted = str(row_vals[start_col]).strip() if start_col < len(row_vals) and row_vals[start_col] is not None else "-"
            correct = str(row_vals[start_col + 1]).strip() if start_col + 1 < len(row_vals) and row_vals[start_col + 1] is not None else "-"
            status = str(row_vals[start_col + 2]).strip() if start_col + 2 < len(row_vals) and row_vals[start_col + 2] is not None else "-"

            for key in ("extracted", "correct", "status"):
                val = locals()[key]
                if val in ("None", ""):
                    locals()[key] = "-"

            # 再取得（locals更新後）
            extracted = "-" if extracted in ("None", "") else extracted
            correct = "-" if correct in ("None", "") else correct
            status = "-" if status in ("None", "") else status

            data[doc_key][field_name] = {
                "extracted": extracted,
                "correct": correct,
                "status": status,
            }

    wb.close()

    doc_types = [key for key, _ in doc_columns]

    return {
        "field_names": field_names,
        "catalog": catalog,
        "doc_types": doc_types,
        "data": data,
    }


# ============================================================
# 2. AB分類
# ============================================================

def classify_ab(parsed: dict) -> tuple[list[str], list[str]]:
    """各項目をA/Bに振り分ける。

    A: カタログに値あり
    B: カタログに値なし（"-" 等）

    Returns:
        (a_fields, b_fields)
    """
    na_values = {"", "-", "ー", "－", "―", "n/a", "na", "none", "None"}
    catalog = parsed["catalog"]

    a_fields: list[str] = []
    b_fields: list[str] = []

    for name in parsed["field_names"]:
        val = catalog.get(name, "-")
        if val in na_values:
            b_fields.append(name)
        else:
            a_fields.append(name)

    logger.info("A項目: %d件, B項目: %d件", len(a_fields), len(b_fields))
    return a_fields, b_fields


# ============================================================
# 3. B項目: 資料間の一致/不一致判定
# ============================================================

def check_b_cross_match(parsed: dict, b_fields: list[str]) -> dict[str, dict]:
    """B項目について、2資料以上に値がある場合に一致/不一致を判定する。

    Returns:
        {項目名: {
            "values": {技術文書キー: 抽出値},
            "cross_status": "一致" / "不一致" / "-"  (1資料以下)
        }}
    """
    na_values = {"", "-", "ー", "－", "―", "不鮮明", "None"}
    data = parsed["data"]
    doc_types = parsed["doc_types"]

    result: dict[str, dict] = {}

    for name in b_fields:
        values: dict[str, str] = {}
        for dt in doc_types:
            ext = data.get(dt, {}).get(name, {}).get("extracted", "-")
            if ext not in na_values:
                values[dt] = ext

        # 一致/不一致判定（2資料以上ある場合）
        if len(values) >= 2:
            unique_vals = set(v.strip().lower() for v in values.values())
            cross_status = "一致" if len(unique_vals) == 1 else "不一致"
        else:
            cross_status = "-"

        result[name] = {"values": values, "cross_status": cross_status}

    return result


# ============================================================
# 4. B項目: AI差分コメント生成
# ============================================================

_B_COMMENT_PROMPT = """\
あなたは照明器具の技術仕様を確認する専門家です。

## タスク1: カタログ外の項目（B項目）
以下の項目はカタログ（SPICE）には記載がなく、他の技術資料から読み取られた値です。
各項目について、どの資料に何が書いてあるかを簡潔に日本語で要約してください。

{b_table}

## タスク2: カタログとの表記揺れ（A項目）
以下の項目はカタログ（SPICE）に記載がありますが、一部の技術資料で表記が異なります。
カタログ値と比較して、どの資料にどのような揺れがあるかを簡潔に説明してください。
完全に同じ表記の資料は言及不要です。揺れがない場合はcommentを空文字にしてください。

{a_table}

## 出力形式
JSON配列で出力してください:
[{{"field": "項目名", "comment": "要約コメント"}}, ...]

コメントは1〜2文で簡潔に。例:
- "器具銘板と取説に 33.5W と記載あり"
- "承認図では「100〜242V」、取説では「AC100V, AC200V, AC242V」と表記が異なる"
- "全資料で同一表記"（→ この場合 comment を空文字に）
"""


def _get_genai_client() -> genai.Client:
    """Gemini APIクライアントを初期化する。"""
    api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        logger.error("環境変数 GOOGLE_API_KEY または GEMINI_API_KEY が未設定です。")
        sys.exit(1)
    return genai.Client(api_key=api_key, http_options={"api_version": "v1alpha"})


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, min=4, max=30),
    reraise=True,
)
def _call_gemini(client: genai.Client, prompt: str) -> str:
    """Gemini APIを呼び出す。"""
    response = client.models.generate_content(model=_MODEL, contents=[prompt])
    text = ""
    if response.candidates:
        for part in response.candidates[0].content.parts:
            if hasattr(part, "text") and part.text:
                text += part.text
    return text


def _extract_json(text: str) -> list[dict] | None:
    """テキストからJSON配列を抽出する。"""
    cleaned = text.strip()
    if "```json" in cleaned:
        cleaned = cleaned.split("```json", 1)[1].split("```", 1)[0].strip()
    elif "```" in cleaned:
        cleaned = cleaned.split("```", 1)[1].split("```", 1)[0].strip()
    if cleaned.startswith("[") or cleaned.startswith("{"):
        try:
            parsed = json.loads(cleaned)
            if isinstance(parsed, dict):
                return [parsed]
            return parsed
        except json.JSONDecodeError:
            pass
    return None


def _find_a_variations(parsed: dict, a_fields: list[str]) -> dict[str, dict[str, str]]:
    """A項目について、カタログと表記が異なる資料を検出する。

    Returns:
        {項目名: {技術文書キー: 抽出値}} — 揺れがある資料のみ
    """
    na_values = {"", "-", "ー", "－", "―", "不鮮明", "None"}
    catalog = parsed["catalog"]
    data = parsed["data"]
    doc_types = parsed["doc_types"]

    result: dict[str, dict[str, str]] = {}
    for name in a_fields:
        cat_val = catalog.get(name, "-")
        if cat_val in na_values:
            continue
        cat_norm = cat_val.strip().lower().replace(" ", "")

        diffs: dict[str, str] = {}
        for dt in doc_types:
            ext = data.get(dt, {}).get(name, {}).get("extracted", "-")
            if ext in na_values:
                continue
            ext_norm = ext.strip().lower().replace(" ", "")
            if ext_norm != cat_norm:
                diffs[dt] = ext

        if diffs:
            result[name] = diffs

    return result


def generate_b_comments(
    b_cross: dict[str, dict],
    parsed: dict | None = None,
    a_fields: list[str] | None = None,
) -> dict[str, str]:
    """B項目の差分コメント + A項目の表記揺れコメントを生成する。

    Returns:
        {項目名: コメント文字列}
    """
    # B項目: コメント対象（値が1つ以上ある項目のみ）
    b_targets = {name: info for name, info in b_cross.items() if info["values"]}

    # A項目: 表記揺れがある項目
    a_variations: dict[str, dict[str, str]] = {}
    if parsed and a_fields:
        a_variations = _find_a_variations(parsed, a_fields)

    if not b_targets and not a_variations:
        return {}

    # B項目テーブル
    b_lines = ["| No. | 項目名 | 記載資料と値 |", "|-----|--------|------------|"]
    if b_targets:
        for i, (name, info) in enumerate(b_targets.items(), 1):
            docs = ", ".join(
                f"{_KEY_TO_DISPLAY.get(dt, dt)}: {val}"
                for dt, val in info["values"].items()
            )
            b_lines.append(f"| {i} | {name} | {docs} |")
    else:
        b_lines.append("| - | (該当なし) | - |")
    b_table = "\n".join(b_lines)

    # A項目テーブル
    catalog = parsed["catalog"] if parsed else {}
    a_lines = ["| No. | 項目名 | カタログ値 | 揺れのある資料と値 |", "|-----|--------|----------|-----------------|"]
    if a_variations:
        for i, (name, diffs) in enumerate(a_variations.items(), 1):
            cat_val = catalog.get(name, "-")
            docs = ", ".join(
                f"{_KEY_TO_DISPLAY.get(dt, dt)}: {val}"
                for dt, val in diffs.items()
            )
            a_lines.append(f"| {i} | {name} | {cat_val} | {docs} |")
    else:
        a_lines.append("| - | (該当なし) | - | - |")
    a_table = "\n".join(a_lines)

    prompt = _B_COMMENT_PROMPT.format(b_table=b_table, a_table=a_table)

    try:
        client = _get_genai_client()
        response_text = _call_gemini(client, prompt)
        results = _extract_json(response_text)
        if not results:
            logger.warning("差分コメントのJSON抽出に失敗: %s", response_text[:200])
            return {}

        comments: dict[str, str] = {}
        for item in results:
            field = item.get("field", "")
            comment = item.get("comment", "")
            if field and comment:
                comments[field] = comment

        logger.info("差分コメント生成: %d / %d 件", len(comments), len(b_targets) + len(a_variations))
        return comments

    except Exception:
        logger.exception("差分コメント生成に失敗")
        return {}


# ============================================================
# 5. Excel出力
# ============================================================

def _write_ab_sheet(
    ws,
    parsed: dict,
    a_fields: list[str],
    b_fields: list[str],
    b_cross: dict[str, dict],
    b_comments: dict[str, str],
) -> None:
    """AB分類を1枚のシートに書き込む。

    レイアウト:
    | 項目名 | SPICEまたはそれに類するもの | 図面(組立図) | 承認図 | ... | B |

    全項目: 各文書列に抽出結果をそのまま表示（記載なしは「-」）
    A項目: カタログと表記揺れがある場合はB列にコメント
    B項目: B列にAI差分コメント
    """
    fill_header = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    fill_catalog = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    fill_b = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap = Alignment(vertical="center", wrap_text=True)

    doc_types = parsed["doc_types"]
    catalog = parsed["catalog"]
    data = parsed["data"]
    na_values = {"", "-", "ー", "－", "―", "不鮮明", "None"}

    # ヘッダー
    headers = ["項目名", "SPICEまたはそれに類するもの"]
    for dt in doc_types:
        headers.append(_KEY_TO_DISPLAY.get(dt, dt))
    headers.append("B")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center
    # B列ヘッダーを黄色に
    ws.cell(row=1, column=len(headers)).fill = fill_b

    b_col = len(headers)
    row_idx = 2

    # A項目
    for name in a_fields:
        ws.cell(row=row_idx, column=1, value=name)

        cat_cell = ws.cell(row=row_idx, column=2, value=catalog.get(name, "-"))
        cat_cell.fill = fill_catalog

        for dt_idx, dt in enumerate(doc_types):
            col = 3 + dt_idx
            info = data.get(dt, {}).get(name, {})
            extracted = info.get("extracted", "-")
            cell = ws.cell(row=row_idx, column=col, value=extracted)
            cell.alignment = wrap

        # B列: 表記揺れコメント
        comment = b_comments.get(name, "")
        b_cell = ws.cell(row=row_idx, column=b_col, value=comment or "-")
        b_cell.alignment = wrap
        if comment:
            b_cell.fill = fill_b

        row_idx += 1

    # B項目
    for name in b_fields:
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value="-")

        for dt_idx, dt in enumerate(doc_types):
            col = 3 + dt_idx
            info = data.get(dt, {}).get(name, {})
            extracted = info.get("extracted", "-")
            cell = ws.cell(row=row_idx, column=col, value=extracted)
            cell.alignment = wrap

        # B列: AI差分コメント
        comment = b_comments.get(name, "")
        if not comment:
            cross_info = b_cross.get(name, {"values": {}})
            if cross_info["values"]:
                comment = "／".join(
                    f"{_KEY_TO_DISPLAY.get(dt, dt)}: {val}"
                    for dt, val in cross_info["values"].items()
                )
        b_cell = ws.cell(row=row_idx, column=b_col, value=comment or "-")
        b_cell.alignment = wrap
        if comment:
            b_cell.fill = fill_b

        row_idx += 1

    # 凡例
    legend_row = row_idx + 1
    ws.cell(row=legend_row, column=1, value="-：記載なし")

    _adjust_widths(ws)


def _adjust_widths(ws) -> None:
    """列幅を自動調整する。"""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
        ):
            for cell in row:
                if cell.value:
                    # 改行がある場合は最長行で計算
                    lines = str(cell.value).split("\n")
                    line_max = max(len(line) for line in lines)
                    if line_max > max_length:
                        max_length = line_max
        ws.column_dimensions[col_letter].width = max(max_length * 1.5 + 2, 8)


def export_ab_excel(
    parsed: dict,
    a_fields: list[str],
    b_fields: list[str],
    b_cross: dict[str, dict],
    b_comments: dict[str, str],
    output_path: str,
) -> None:
    """AB分類結果をExcelに出力する（1シート統合）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AB分類"
    _write_ab_sheet(ws, parsed, a_fields, b_fields, b_cross, b_comments)

    wb.save(output_path)
    wb.close()
    logger.info("AB分類Excel出力: %s", output_path)


# ============================================================
# メイン
# ============================================================

def main():
    if len(sys.argv) < 2:
        print("使い方: uv run python ab_classifier.py <正誤判定Excelパス>")
        sys.exit(1)

    input_path = sys.argv[1]
    if not os.path.exists(input_path):
        print(f"ファイルが見つかりません: {input_path}")
        sys.exit(1)

    # 出力パス: 同じディレクトリに AB分類_品番_タイムスタンプ.xlsx
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    input_dir = os.path.dirname(input_path)
    input_base = os.path.splitext(os.path.basename(input_path))[0]
    # 正誤結果_XXX_YYY → AB分類_XXX_YYY
    output_name = input_base.replace("正誤結果", "AB分類") + f"_{timestamp}.xlsx"
    output_path = os.path.join(input_dir or ".", output_name)

    logger.info("入力: %s", input_path)

    # 1. Excel読み込み
    parsed = parse_comparison_excel(input_path)
    logger.info(
        "読み込み完了: %d項目, %d技術文書",
        len(parsed["field_names"]),
        len(parsed["doc_types"]),
    )

    # 2. AB分類
    a_fields, b_fields = classify_ab(parsed)

    # 3. B項目: 資料間の一致/不一致判定
    b_cross = check_b_cross_match(parsed, b_fields)

    # 4. B項目: AI差分コメント生成
    b_comments = {}
    if b_fields:
        logger.info("B項目の差分コメント + A項目の表記揺れコメントを生成中...")
        b_comments = generate_b_comments(b_cross, parsed, a_fields)

    # 5. Excel出力
    export_ab_excel(parsed, a_fields, b_fields, b_cross, b_comments, output_path)

    print(f"\n出力完了: {output_path}")
    print(f"  A項目: {len(a_fields)}件")
    print(f"  B項目: {len(b_fields)}件")


if __name__ == "__main__":
    main()
