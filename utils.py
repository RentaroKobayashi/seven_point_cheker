"""共通ユーティリティ（CSV/Excel読み込み、色割り当て、Geminiクライアント初期化）"""

import colorsys
import csv
import io
import os
import re
import unicodedata
import zipfile

import fitz  # PyMuPDF
import openpyxl
import streamlit as st
from dotenv import load_dotenv
from google import genai

load_dotenv()

# 20色パレット（視認性の高い色）
COLOR_PALETTE = [
    "#E53935",  # 赤
    "#00897B",  # ティール
    "#039BE5",  # スカイブルー
    "#FB8C00",  # オレンジ
    "#8E24AA",  # 紫
    "#43A047",  # 緑
    "#6D4C41",  # 茶
    "#F06292",  # ピンク
    "#3949AB",  # インディゴ
    "#00ACC1",  # シアン
    "#FFB300",  # アンバー
    "#7CB342",  # ライトグリーン
    "#5E35B1",  # ディープパープル
    "#D81B60",  # ローズ
    "#1E88E5",  # ブルー
    "#C0CA33",  # ライム
    "#546E7A",  # ブルーグレー
    "#FF7043",  # ディープオレンジ
    "#26A69A",  # ミディアムティール
    "#AB47BC",  # ミディアムパープル
]


def assign_colors(fields: list[dict]) -> dict[str, str]:
    """項目リストに動的に色を割り当てる。

    20色パレットを超える場合は黄金角でHSV生成。
    """
    colors: dict[str, str] = {}
    for i, field in enumerate(fields):
        name = field["name"]
        if i < len(COLOR_PALETTE):
            colors[name] = COLOR_PALETTE[i]
        else:
            # 黄金角による均等分散
            hue = (i * 0.618033988749895) % 1.0
            r, g, b = colorsys.hsv_to_rgb(hue, 0.7, 0.85)
            colors[name] = f"#{int(r*255):02X}{int(g*255):02X}{int(b*255):02X}"
    return colors


def load_items_from_file(uploaded_file) -> list[dict]:
    """CSV または Excel ファイルを読み込み、項目リストを返す。

    CSV の場合:
        列1: 項目名（必須）、列2: 説明（任意）、列3: 読み取りヒント（任意）

    Excel の場合（項目確認ドラフト形式に対応）:
        「詳細項目」列 → 項目名
        「成績書項目名」列 → 説明（カテゴリ）
        「具体値」列 → 読み取りヒント
    """
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xls")):
        return _load_items_from_excel(uploaded_file)
    return _load_items_from_csv(uploaded_file)


def _load_items_from_csv(uploaded_file) -> list[dict]:
    """CSVファイルを読み込み、項目リストを返す。"""
    content = uploaded_file.getvalue().decode("utf-8-sig")
    reader = csv.reader(io.StringIO(content))

    fields: list[dict] = []
    for row in reader:
        if not row or not row[0].strip():
            continue
        # ヘッダー行をスキップ
        if row[0].strip() in ("項目名", "name", "項目"):
            continue
        field = {
            "name": row[0].strip(),
            "description": row[1].strip() if len(row) > 1 and row[1].strip() else "",
            "hint": row[2].strip() if len(row) > 2 and row[2].strip() else "",
        }
        fields.append(field)

    if not fields:
        st.error("CSVに有効な項目が含まれていません。1列目に項目名を入力してください。")
        return []

    return fields


# 資料種別列のマッピング（列ヘッダー → カテゴリ名）
_DOC_TYPE_MAP = {
    "図面(組み立て図)": "組立図",
    "図面": "組立図",
    "組立図": "組立図",
    "承認図": "承認図",
    "器具銘板": "器具銘板",
    "銘板": "器具銘板",
    "取説": "取扱説明書",
    "取扱説明書": "取扱説明書",
    "外装表示": "外装ラベル",
    "外装ラベル": "外装ラベル",
}


def _load_items_from_excel(uploaded_file) -> list[dict]:
    """Excelファイル（項目確認ドラフト形式）を読み込み、項目リストを返す。

    ヘッダー行を自動検出し、「詳細項目」列を項目名として使用する。
    """
    wb = openpyxl.load_workbook(io.BytesIO(uploaded_file.getvalue()), read_only=True)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if not rows:
        st.error("Excelファイルにデータがありません。")
        return []

    # ヘッダー行を自動検出（「詳細項目」を含む行）
    header_idx = None
    headers: list[str] = []
    for i, row in enumerate(rows):
        row_strs = [str(c).strip() if c else "" for c in row]
        if "詳細項目" in row_strs:
            header_idx = i
            headers = row_strs
            break

    if header_idx is None:
        st.error("Excelに「詳細項目」列が見つかりません。")
        return []

    # 列インデックスを特定
    def _find_col(candidates: list[str]) -> int | None:
        for name in candidates:
            if name in headers:
                return headers.index(name)
        return None

    col_name = _find_col(["詳細項目"])
    col_desc = _find_col(["成績書項目名"])
    col_hint = _find_col(["具体値"])

    if col_name is None:
        st.error("「詳細項目」列が見つかりません。")
        return []

    # 資料種別の列を検出
    doc_type_cols: dict[str, int] = {}
    for idx, h in enumerate(headers):
        mapped = _DOC_TYPE_MAP.get(h)
        if mapped:
            doc_type_cols[mapped] = idx

    # データ行を読み込み
    fields: list[dict] = []
    seen_names: set[str] = set()

    for row in rows[header_idx + 1 :]:
        cells = list(row)
        # 項目名を取得
        raw_name = str(cells[col_name]).strip() if col_name < len(cells) and cells[col_name] else ""
        if not raw_name or raw_name == "None":
            continue
        # 番兵行（1文字のダミー値等）をスキップ
        if len(raw_name) <= 1 and not raw_name.isdigit():
            continue

        # 重複スキップ
        if raw_name in seen_names:
            continue
        seen_names.add(raw_name)

        # 説明
        desc = ""
        if col_desc is not None and col_desc < len(cells) and cells[col_desc]:
            desc = str(cells[col_desc]).strip()
            if desc == "None":
                desc = ""

        # ヒント（具体値）
        hint = ""
        if col_hint is not None and col_hint < len(cells) and cells[col_hint]:
            hint = str(cells[col_hint]).strip()
            if hint == "None":
                hint = ""

        # どの資料種別で照合するか
        doc_types: list[str] = []
        for doc_name, col_idx in doc_type_cols.items():
            if col_idx < len(cells) and cells[col_idx]:
                val = str(cells[col_idx]).strip()
                if val and val != "ー" and val != "-" and val != "None":
                    doc_types.append(doc_name)

        field: dict = {
            "name": raw_name,
            "description": desc,
            "hint": hint,
        }
        if doc_types:
            field["doc_types"] = doc_types

        fields.append(field)

    if not fields:
        st.error("Excelに有効な項目が含まれていません。")
        return []

    return fields


def pdf_to_images(pdf_bytes: bytes, dpi: int = 200) -> list[dict]:
    """PDFバイトデータを各ページのPNG画像に変換する。

    Returns:
        [{"name": "ファイル名_p1.png", "data": bytes, "mime_type": "image/png"}, ...]
    """
    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    images: list[dict] = []
    zoom = dpi / 72  # 72dpi が fitz のデフォルト
    mat = fitz.Matrix(zoom, zoom)

    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(matrix=mat)
        png_data = pix.tobytes("png")
        images.append(
            {
                "data": png_data,
                "mime_type": "image/png",
                "page_num": page_num + 1,
            }
        )

    doc.close()
    return images


# ファイル名から品番を抽出するためのカテゴリキーワード
_PRODUCT_KEYWORDS = re.compile(
    r"カタログ|catalog|承認図|組立図|組み立て図|器具銘板|銘板"
    r"|取扱説明|取説|説明書|外装ラベル|外装表示|外装|図面",
    re.IGNORECASE,
)


def expand_zip_file(zip_data: bytes) -> list[dict]:
    """ZIPを展開し、画像ファイルのリストを返す。

    ZIPエントリのパスからフォルダ名を product_id として付与する。
    隠しファイル（__MACOSX、.始まり）はスキップ。

    Returns:
        [{"name": str, "data": bytes, "mime_type": str, "product_id": str}, ...]
    """
    expanded: list[dict] = []

    with zipfile.ZipFile(io.BytesIO(zip_data)) as zf:
        for entry in zf.namelist():
            # 隠しファイル・ディレクトリをスキップ
            if "__MACOSX" in entry:
                continue
            parts = entry.split("/")
            if any(p.startswith(".") for p in parts):
                continue
            # ディレクトリエントリはスキップ
            if entry.endswith("/"):
                continue

            basename = os.path.basename(entry)
            lower = basename.lower()

            # 画像・PDF以外はスキップ
            if not lower.endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp", ".pdf")):
                continue

            data = zf.read(entry)

            # フォルダ名から product_id を取得（ルート直下のフォルダ名）
            path_parts = entry.split("/")
            if len(path_parts) >= 2 and path_parts[0]:
                product_id = path_parts[0]
            else:
                product_id = ""

            if lower.endswith(".pdf"):
                # PDFをページごとの画像に変換
                pages = pdf_to_images(data)
                base = os.path.splitext(basename)[0]
                if len(pages) == 1:
                    expanded.append(
                        {
                            "name": f"{base}.png",
                            "data": pages[0]["data"],
                            "mime_type": "image/png",
                            "product_id": product_id,
                        }
                    )
                else:
                    for page in pages:
                        expanded.append(
                            {
                                "name": f"{base}_p{page['page_num']}.png",
                                "data": page["data"],
                                "mime_type": "image/png",
                                "product_id": product_id,
                            }
                        )
            else:
                mime = "image/png" if lower.endswith(".png") else "image/jpeg"
                expanded.append(
                    {
                        "name": basename,
                        "data": data,
                        "mime_type": mime,
                        "product_id": product_id,
                    }
                )

    return expanded


def extract_product_from_filename(filename: str) -> str | None:
    """ファイル名から品番を抽出する。

    カテゴリキーワード（カタログ、組立図等）の前の _ で分割して品番部分を取得する。
    キーワードが見つからない or 品番部分が空の場合は None を返す。

    例:
        "NEL4600_カタログ.jpg" → "NEL4600"
        "NTS32711S_取説_p1.png" → "NTS32711S"
        "3_NEL4600ENRZ9承認図.png" → "NEL4600ENRZ9"
    """
    # 拡張子を除去（NFC正規化でmacOSのNFD濁点問題を回避）
    base = os.path.splitext(unicodedata.normalize("NFC", filename))[0]

    # キーワードの位置を検出
    match = _PRODUCT_KEYWORDS.search(base)
    if not match:
        return None

    # キーワードの前の部分を取得
    before = base[: match.start()]

    # 末尾の区切り文字（_ - スペース）を除去
    before = before.rstrip("_- ")

    if not before:
        return None

    # _ で分割し、先頭の数字のみの部分をスキップして品番を取得
    # 例: "3_NEL4600ENRZ9" → ["3", "NEL4600ENRZ9"] → "NEL4600ENRZ9"
    parts = [p.strip() for p in before.split("_") if p.strip()]
    # 数字のみの部分（連番）をスキップ
    non_numeric = [p for p in parts if not p.isdigit()]
    if non_numeric:
        return non_numeric[0]
    # 全て数字の場合は最初の部分を返す（フォールバック）
    return parts[0] if parts else None


def extract_product_from_items_filename(filename: str) -> str | None:
    """項目リストファイル名から品番を抽出する。

    ファイル名パターン `{品番}_項目リスト.xlsx` から品番部分を取得する。
    パターンにマッチしない場合は None を返す。

    例:
        "NEL4600ENRZ9_項目リスト.xlsx" → "NEL4600ENRZ9"
        "NTS32711S_項目リスト.xlsx" → "NTS32711S"
        "SQ-LD440-K_項目リスト.xlsx" → "SQ-LD440-K"
    """
    base = os.path.splitext(filename)[0]
    m = re.match(r"^(.+?)_項目リスト$", base)
    if m:
        return m.group(1)
    return None


def group_images_by_product(expanded: list[dict]) -> dict[str, list[dict]]:
    """画像リストを品番ごとにグルーピングする。

    優先順位:
        1. product_id フィールド（ZIP展開時にフォルダ名から付与）が非空ならそれを使う
        2. ファイル名から extract_product_from_filename() で品番抽出
        3. いずれも該当しない → "_default" グループ

    Returns:
        {"品番1": [画像リスト], "品番2": [画像リスト], "_default": [...]}
    """
    groups: dict[str, list[dict]] = {}

    for img in expanded:
        # 1. product_id フィールドを優先
        product_id = img.get("product_id", "")
        if product_id:
            groups.setdefault(product_id, []).append(img)
            continue

        # 2. ファイル名から品番抽出
        product = extract_product_from_filename(img["name"])
        if product:
            groups.setdefault(product, []).append(img)
            continue

        # 3. いずれも該当しない
        groups.setdefault("_default", []).append(img)

    return groups


def expand_uploaded_files(uploaded_files: list) -> list[dict]:
    """アップロードファイルを展開する。PDFは各ページを画像に変換、ZIPは展開する。

    Returns:
        [{"name": str, "data": bytes, "mime_type": str, "product_id": str}, ...]
    """
    expanded: list[dict] = []

    for f in uploaded_files:
        name = f.name
        data = f.getvalue()

        if name.lower().endswith(".zip"):
            # ZIP展開（product_id はフォルダ名から自動付与）
            expanded.extend(expand_zip_file(data))
        elif name.lower().endswith(".pdf"):
            pages = pdf_to_images(data)
            base = os.path.splitext(name)[0]
            if len(pages) == 1:
                expanded.append(
                    {
                        "name": f"{base}.png",
                        "data": pages[0]["data"],
                        "mime_type": "image/png",
                        "product_id": "",
                    }
                )
            else:
                for page in pages:
                    expanded.append(
                        {
                            "name": f"{base}_p{page['page_num']}.png",
                            "data": page["data"],
                            "mime_type": "image/png",
                            "product_id": "",
                        }
                    )
        else:
            mime = "image/png" if name.lower().endswith(".png") else "image/jpeg"
            expanded.append({"name": name, "data": data, "mime_type": mime, "product_id": ""})

    return expanded


def get_genai_client() -> genai.Client:
    """Gemini APIクライアントを初期化して返す。"""
    api_key = os.environ.get("GOOGLE_API_KEY") or os.environ.get("GEMINI_API_KEY")
    if not api_key:
        st.error(
            "環境変数 `GOOGLE_API_KEY` または `GEMINI_API_KEY` が設定されていません。"
            "`.env` ファイルを確認してください。"
        )
        st.stop()
    return genai.Client(api_key=api_key, http_options={"api_version": "v1alpha"})


def load_master_excel(file) -> tuple[list[str], dict[str, list[str]]]:
    """マスターExcelの「詳細項目」シートから品番一覧と品番別チェック項目を取得する。

    「詳細項目」シートの構造:
        Row 1: 番号ヘッダー（None, 1, 2, 3, ..., 15）→スキップ
        Row 2以降: Col A = 品番, Col B〜P = チェック項目名

    Returns:
        (品番リスト, {品番: [項目名リスト]})
    """
    wb = openpyxl.load_workbook(io.BytesIO(file.getvalue()), read_only=True)
    ws = wb["詳細項目"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    product_numbers: list[str] = []
    items_by_product: dict[str, list[str]] = {}

    # Row 1（index 0）は番号ヘッダーなのでスキップし、Row 2以降を処理
    for row in rows[1:]:
        product_number = row[0] if row else None
        if product_number is None or str(product_number).strip() == "":
            continue
        product_number = str(product_number).strip()

        # Col B〜P（index 1〜15）からチェック項目名を取得
        item_names: list[str] = []
        for cell in row[1:16]:
            if cell is not None and str(cell).strip() != "":
                item_names.append(str(cell).strip())

        product_numbers.append(product_number)
        items_by_product[product_number] = item_names

    return product_numbers, items_by_product


def _parse_multi_header_rows(
    rows: list[tuple],
) -> dict[str, dict[str, dict[str, str]]]:
    """品番ごとにヘッダー行を持つ形式を解析する。

    各品番セクション:
        ヘッダー行: Col A=品番, Col B=技術文書, Col C以降=項目名（品番ごとに異なる）
        データ行: Col A=品番値, Col B=技術文書種別, Col C以降=正解値
        空行で区切り

    Returns:
        {品番: {技術文書: {項目名: 正解値}}}
    """
    result: dict[str, dict[str, dict[str, str]]] = {}
    current_items: list[str] = []
    i = 0

    while i < len(rows):
        row = rows[i]
        cells = [str(c).strip() if c is not None else "" for c in row]

        # 空行スキップ
        if not any(cells):
            i += 1
            continue

        # ヘッダー行の検出（Col A="品番" かつ Col B="技術文書"）
        if len(cells) >= 2 and cells[0] == "品番" and cells[1] == "技術文書":
            current_items = [c for c in cells[2:] if c]
            i += 1
            continue

        # データ行（ヘッダーが設定済みの場合）
        if current_items and len(cells) >= 2 and cells[0] and cells[1]:
            product = cells[0]
            doc_type = cells[1]
            values: dict[str, str] = {}
            for j, name in enumerate(current_items):
                cell_idx = 2 + j
                val = cells[cell_idx] if cell_idx < len(cells) else ""
                values[name] = val

            result.setdefault(product, {})[doc_type] = values

        i += 1

    return result


def _parse_single_header_rows(
    rows: list[tuple],
) -> dict[str, dict[str, dict[str, str]]]:
    """従来形式（全品番共通ヘッダー）を解析する。

    Row 1: 空行
    Row 2: ヘッダー（Col B=品番, Col C=技術文書, Col D以降=項目名）
    Row 3以降: データ行

    Returns:
        {品番: {技術文書: {項目名: 正解値}}}
    """
    if len(rows) < 3:
        return {}

    header_row = rows[1]
    item_names: list[str] = []
    for cell in header_row[3:]:
        item_names.append(str(cell).strip() if cell is not None else "")

    result: dict[str, dict[str, dict[str, str]]] = {}

    for row in rows[2:]:
        row_product = row[1] if len(row) > 1 else None
        if row_product is None or str(row_product).strip() == "":
            continue
        product = str(row_product).strip()

        doc_type = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ""
        if not doc_type:
            continue

        values: dict[str, str] = {}
        for i, name in enumerate(item_names):
            if not name:
                continue
            cell_idx = 3 + i
            cell_val = row[cell_idx] if len(row) > cell_idx else None
            values[name] = str(cell_val).strip() if cell_val is not None else ""

        result.setdefault(product, {})[doc_type] = values

    return result


def _detect_format(rows: list[tuple]) -> str:
    """Excelの形式を自動判定する。

    Returns:
        "multi_header": 品番ごとにヘッダー行がある形式
        "single_header": 従来の全品番共通ヘッダー形式
    """
    for row in rows:
        cells = [str(c).strip() if c is not None else "" for c in row]
        if len(cells) >= 2 and cells[0] == "品番" and cells[1] == "技術文書":
            return "multi_header"
    return "single_header"


def load_correct_data_from_master(
    file, product_number: str
) -> dict[str, dict[str, str]]:
    """マスターExcelのSheet1から指定品番の正解データを取得する。

    品番ごとにヘッダーが異なる形式と、従来の共通ヘッダー形式の両方に対応。

    Returns:
        {技術文書: {項目名: 正解値}}
    """
    if isinstance(file, str):
        wb = openpyxl.load_workbook(file, read_only=True)
    else:
        wb = openpyxl.load_workbook(io.BytesIO(file.getvalue()), read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    all_data = _detect_and_parse(rows)
    return all_data.get(product_number, {})


def load_all_correct_data_from_master(
    file_path: str,
) -> dict[str, dict[str, dict[str, str]]]:
    """マスターExcelのSheet1から全品番の正解データを取得する。

    品番ごとにヘッダーが異なる形式と、従来の共通ヘッダー形式の両方に対応。

    Returns:
        {品番: {技術文書: {項目名: 正解値}}}
    """
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    return _detect_and_parse(rows)


def _detect_and_parse(
    rows: list[tuple],
) -> dict[str, dict[str, dict[str, str]]]:
    """形式を自動判定して解析する。"""
    fmt = _detect_format(rows)
    if fmt == "multi_header":
        return _parse_multi_header_rows(rows)
    return _parse_single_header_rows(rows)
