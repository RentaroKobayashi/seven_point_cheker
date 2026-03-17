"""正解データ比較 + 正答率計算"""

import io
import json
import logging
import os
import re

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from tenacity import retry, stop_after_attempt, wait_exponential

logger = logging.getLogger(__name__)


def load_correct_data(excel_path: str) -> tuple[list[str], dict]:
    """正解ExcelをロードしてDict化する。

    Args:
        excel_path: Excelファイルパス

    Returns:
        (ヘッダーリスト, {ソース名: [{列名: 値, ...}, ...]})
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # シート名「正解データ」を探す。なければ最初のシート
    if "正解データ" in wb.sheetnames:
        sheet = wb["正解データ"]
    else:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return [], {}

    headers = [str(h).strip() if h else "" for h in rows[0]]
    data: dict[str, list[dict]] = {}

    for row in rows[1:]:
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                row_dict[headers[i]] = str(val).strip() if val is not None else ""

        # ソース列を特定（「ソース」列があればそれを使用）
        source = row_dict.get("ソース", "")
        if not source:
            # 最初の列をソースとして使用
            source = str(row[0]).strip() if row[0] else ""

        if source:
            data.setdefault(source, []).append(row_dict)

    return headers, data


def normalize_for_comparison(value: str) -> str:
    """比較用に値を正規化する。"""
    if not value:
        return ""

    s = value.strip()

    # 全角英数字→半角
    s = s.translate(
        str.maketrans(
            "ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ０１２３４５６７８９",
            "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789",
        )
    )

    # 複数スペースを単一に
    s = re.sub(r"[\s　]+", " ", s)
    s = s.strip()

    return s


def _extract_all_numbers(s: str) -> list[float]:
    """文字列から全ての数値を抽出する（出現順）。"""
    return [float(x) for x in re.findall(r"[\d]+(?:\.[\d]+)?", s)]


def _extract_range(s: str) -> tuple[float, float] | None:
    """範囲表記（100~242, 100-242, 100〜242 等）を検出する。"""
    m = re.search(
        r"([\d]+(?:\.[\d]+)?)\s*[~～〜\-ー−]\s*([\d]+(?:\.[\d]+)?)", s
    )
    if m:
        lo, hi = float(m.group(1)), float(m.group(2))
        return (min(lo, hi), max(lo, hi))
    return None


def _strip_decorations(s: str) -> str:
    """AC/DC等の接頭辞、括弧表記の揺れ、記号の揺れを除去して正規化する。"""
    s = re.sub(r"(?i)\b(AC|DC)\s*", "", s)
    # 全角括弧→半角
    s = s.replace("（", "(").replace("）", ")")
    # 中点・カンマ等の区切りを統一
    s = re.sub(r"[・、，,]\s*", ",", s)
    return s.strip()


def is_numeric_match(extracted: str, correct: str) -> bool | None:
    """数値比較。一致ならTrue、不一致ならFalse、比較不能ならNone。"""
    ext_nums = _extract_all_numbers(extracted)
    cor_nums = _extract_all_numbers(correct)

    if not ext_nums and not cor_nums:
        return None
    if not ext_nums or not cor_nums:
        return False

    ext_set = set(ext_nums)
    cor_set = set(cor_nums)

    # 数値セットが完全一致
    if ext_set == cor_set:
        return True

    # 範囲 vs リスト の比較
    ext_range = _extract_range(extracted)
    cor_range = _extract_range(correct)

    if cor_range and ext_nums:
        lo, hi = cor_range
        if all(lo <= n <= hi for n in ext_nums):
            return True

    if ext_range and cor_nums:
        lo, hi = ext_range
        if all(lo <= n <= hi for n in cor_nums):
            return True

    # 数値が1つだけ同士の比較
    if len(ext_nums) == 1 and len(cor_nums) == 1:
        return ext_nums[0] == cor_nums[0]

    return None


def _numeric_overlap_ratio(ext: str, cor: str) -> float:
    """2つの文字列に含まれる数値セットの重複率（0.0〜1.0）。"""
    ext_set = set(_extract_all_numbers(ext))
    cor_set = set(_extract_all_numbers(cor))
    if not ext_set or not cor_set:
        return 0.0
    intersection = ext_set & cor_set
    union = ext_set | cor_set
    return len(intersection) / len(union) if union else 0.0


def determine_match_status(extracted: str, correct: str, field_name: str) -> str:
    """比較判定を行う。

    判定フロー:
    1. 両方N/A → "na"
    2. 正規化後の完全一致 → "match"
    3. 数値セット比較（範囲・リスト対応） → "match"
    4. 接頭辞除去後の一致 → "match"
    5. 部分文字列一致 → "mismatch"
    6. 数値の部分的重複 → "mismatch"
    7. いずれにも該当しない → "mismatch"

    Returns:
        "match" / "mismatch" / "na"
    """
    na_values = {"", "-", "ー", "－", "―", "n/a", "na", "none"}
    ext_norm = normalize_for_comparison(extracted)
    cor_norm = normalize_for_comparison(correct)
    ext_lower = ext_norm.lower()
    cor_lower = cor_norm.lower()

    # 1. N/A判定
    ext_na = ext_lower in na_values
    cor_na = cor_lower in na_values
    if ext_na and cor_na:
        return "na"
    if cor_na and not ext_na:
        # 正解データになく抽出できた → 対象外扱い（OKとする）
        return "na"
    if ext_na and not cor_na:
        # 正解データにあるが抽出できなかった → NG
        return "mismatch"

    # 2. 正規化後の完全一致
    if ext_lower == cor_lower:
        return "match"

    # スペース・記号除去して一致
    ext_clean = ext_lower.replace(" ", "")
    cor_clean = cor_lower.replace(" ", "")
    if ext_clean == cor_clean:
        return "match"

    # 3. 数値セット比較（範囲 vs リスト にも対応）
    num_result = is_numeric_match(ext_norm, cor_norm)
    if num_result is True:
        return "match"

    # 4. 接頭辞・装飾除去後の比較
    ext_stripped = _strip_decorations(ext_lower)
    cor_stripped = _strip_decorations(cor_lower)
    if ext_stripped and cor_stripped:
        if ext_stripped == cor_stripped:
            return "match"
        # 装飾除去後の数値比較
        stripped_num = is_numeric_match(ext_stripped, cor_stripped)
        if stripped_num is True:
            return "match"

    # 5. 部分文字列一致
    if ext_lower in cor_lower or cor_lower in ext_lower:
        return "mismatch"
    if ext_stripped and cor_stripped:
        if ext_stripped in cor_stripped or cor_stripped in ext_stripped:
            return "mismatch"

    # 6. 数値の部分的重複
    overlap = _numeric_overlap_ratio(ext_norm, cor_norm)
    if overlap >= 0.5:
        return "mismatch"

    return "mismatch"


def compare_with_correct_data(
    results: dict, correct_data: dict, fields: list[dict]
) -> dict:
    """全抽出結果と正解データを一括比較する。

    Args:
        results: {画像名: {"data": [...]}} - 抽出結果
        correct_data: {ソース名: [{列名: 値}]} - 正解データ
        fields: [{name, description, hint}] - 項目リスト

    Returns:
        {画像名: {品番: {項目名: {"extracted": ..., "correct": ..., "status": ...}}}}
    """
    field_names = [f["name"] for f in fields]
    correct_headers = set()
    for entries in correct_data.values():
        for entry in entries:
            correct_headers.update(entry.keys())

    comparison: dict = {}

    for img_name, result in results.items():
        data_list = result.get("data", [])
        if not data_list:
            continue

        img_comparison: dict = {}
        source_key = os.path.splitext(img_name)[0]

        # 正解データのソースキーを柔軟にマッチ
        matched_correct = correct_data.get(source_key, [])
        if not matched_correct:
            # 部分一致で探す
            for ck, cv in correct_data.items():
                if ck in source_key or source_key in ck:
                    matched_correct = cv
                    break

        for item in data_list:
            first_field_val = item.get(field_names[0], "-") if field_names else "-"
            item_key = first_field_val

            item_comparison: dict = {}
            # 対応する正解エントリを探す
            correct_entry = {}
            if matched_correct:
                correct_entry = matched_correct[0] if len(matched_correct) == 1 else {}
                # 品番等で一致するエントリを探す
                for ce in matched_correct:
                    if ce.get(field_names[0], "") == first_field_val:
                        correct_entry = ce
                        break

            for name in field_names:
                extracted_val = item.get(name, "-") or "-"
                correct_val = correct_entry.get(name, "") if correct_entry else ""

                status = determine_match_status(extracted_val, correct_val, name)
                item_comparison[name] = {
                    "extracted": extracted_val,
                    "correct": correct_val,
                    "status": status,
                }

            img_comparison[item_key] = item_comparison

        comparison[img_name] = img_comparison

    return comparison


# 分類カテゴリ → 正解データの技術文書キーへのマッピング
DOC_TYPE_MAP = {
    "組立図": "図面",
    "図面": "図面",
    "承認図": "承認図",
    "器具銘板": "器具銘板",
    "取扱説明書": "取扱説明書",
    "外装ラベル": "外装ラベル",
}

# 正解データ側の表記ゆれを吸収するマッピング（正解データキー → 正規キー）
_CORRECT_KEY_ALIASES = {
    "取説": "取扱説明書",
}


def _similarity_to_hint(candidate: str, hint: str) -> int:
    """候補値とカタログヒントの類似度スコアを返す（高いほど近い）。

    Returns:
        3: 正規化後の完全一致
        2: 部分一致（片方が他方を含む）
        1: 数値部分が一致
        0: 不一致
    """
    c_norm = normalize_for_comparison(candidate)
    h_norm = normalize_for_comparison(hint)

    if not c_norm or not h_norm:
        return 0

    # 完全一致
    if c_norm.lower() == h_norm.lower():
        return 3

    # スペース除去して一致
    if c_norm.lower().replace(" ", "") == h_norm.lower().replace(" ", ""):
        return 3

    # 部分一致
    if c_norm.lower() in h_norm.lower() or h_norm.lower() in c_norm.lower():
        return 2

    # 数値一致
    num_result = is_numeric_match(c_norm, h_norm)
    if num_result is True:
        return 1

    return 0


def _resolve_conflict(candidates: list[str], catalog_hint: str) -> str:
    """複数の候補値からカタログヒントに最も近いものを選ぶ。

    Args:
        candidates: 有効な候補値リスト（"-"や"不鮮明"を除いた値）
        catalog_hint: カタログから読み取った参考値

    Returns:
        採用する値
    """
    if len(candidates) == 1:
        return candidates[0]

    # カタログヒントがない場合 → 全候補を併記
    if not catalog_hint or catalog_hint == "-":
        unique = list(dict.fromkeys(candidates))  # 順序保持で重複除去
        if len(unique) == 1:
            return unique[0]
        return "／".join(unique)

    # 各候補のスコアを計算し、最高スコアの値を返す
    best_score = -1
    best_val = candidates[0]
    for val in candidates:
        score = _similarity_to_hint(val, catalog_hint)
        if score > best_score:
            best_score = score
            best_val = val

    return best_val


_CONFLICT_RESOLUTION_PROMPT = """\
あなたは照明器具の技術仕様値を統合する専門家です。
同じ技術文書の複数ページから、同じ項目に対して異なる値が読み取られました。
各項目について、最も適切な候補を1つ選んでください。

## 選択基準
- カタログ参考値がある場合、それと整合性のある候補を優先
- その項目にとって本来の情報を含む候補を優先
  （例: 「設置要件」なら施工・設置に関する注意、
        「本体材質」なら材料名の列挙）
- 明らかに別の項目の内容を誤って拾ったものは除外
- 情報量がより多い（抜けが少ない）候補を優先

## 判定対象
{table}

JSON配列で出力: [{{"index": 1, "selected_index": 0}}, ...]
selected_index は candidates 配列の0始まりインデックス。
"""

_CONFLICT_MODEL = "gemini-3-flash-preview"


@retry(
    stop=stop_after_attempt(3),
    wait=wait_exponential(multiplier=2, min=4, max=30),
    reraise=True,
)
def _call_gemini_for_conflict(client, prompt: str) -> str:
    """コンフリクト解決用にGemini APIを呼び出す。"""
    response = client.models.generate_content(
        model=_CONFLICT_MODEL,
        contents=[prompt],
    )
    text = ""
    if response.candidates:
        for part in response.candidates[0].content.parts:
            if hasattr(part, "text") and part.text:
                text += part.text
    return text


def _resolve_conflict_with_llm(
    client,
    conflicts: list[dict],
) -> dict[int, int]:
    """Geminiにコンフリクト候補を渡し、最適な候補のインデックスを選ばせる。

    Args:
        client: google.genai.Client インスタンス
        conflicts: [{"index": int, "field": str, "doc_type": str,
                      "candidates": list[str], "catalog_hint": str}]

    Returns:
        {conflict_index: selected_candidate_index} の辞書。
        失敗した項目は含まれない。
    """
    if not conflicts:
        return {}

    # テーブルを構築
    lines = [
        "| No. | 項目名 | 技術文書 | カタログ参考値 | 候補 |",
        "|-----|--------|----------|--------------|------|",
    ]
    for c in conflicts:
        candidates_str = " / ".join(
            f"[{i}] {v}" for i, v in enumerate(c["candidates"])
        )
        hint = c.get("catalog_hint", "-") or "-"
        lines.append(
            f"| {c['index']} | {c['field']} | {c['doc_type']} | {hint} | {candidates_str} |"
        )
    table = "\n".join(lines)
    prompt = _CONFLICT_RESOLUTION_PROMPT.format(table=table)

    try:
        response_text = _call_gemini_for_conflict(client, prompt)

        # JSONを抽出
        cleaned = response_text.strip()
        if "```json" in cleaned:
            cleaned = cleaned.split("```json", 1)[1].split("```", 1)[0].strip()
        elif "```" in cleaned:
            cleaned = cleaned.split("```", 1)[1].split("```", 1)[0].strip()

        results: list[dict] | None = None
        if cleaned.startswith("[") or cleaned.startswith("{"):
            try:
                parsed = json.loads(cleaned)
                if isinstance(parsed, dict):
                    results = [parsed]
                else:
                    results = parsed
            except json.JSONDecodeError:
                pass

        if not results:
            logger.warning("コンフリクトLLM解決: JSONの抽出に失敗。レスポンス: %s", response_text[:200])
            return {}

        selections: dict[int, int] = {}
        for item in results:
            idx = item.get("index")
            sel = item.get("selected_index")
            if idx is not None and sel is not None:
                selections[int(idx)] = int(sel)

        logger.info("コンフリクトLLM解決: %d / %d 件の判定を取得", len(selections), len(conflicts))
        return selections

    except Exception:
        logger.exception("コンフリクトLLM解決: Gemini API呼び出しに失敗")
        return {}


def _merge_by_doc_type(
    results: dict,
    classification: dict,
    fields: list[dict],
    catalog_data: dict | None = None,
    client=None,
) -> dict[str, dict]:
    """同じ技術文書タイプの抽出結果を統合する。

    例: 取扱説明書が3ページ(3画像)ある場合 → 1つの取扱説明書データに統合。

    統合ルール:
    1. 値が1つだけ → そのまま採用
    2. 全ページ同じ値 → そのまま採用
    3. 複数ページで異なる値（コンフリクト） → LLM判定（clientがある場合）
    4. LLMなし or 失敗時 → カタログ値と比較して最も近いものを採用

    Args:
        catalog_data: カタログから抽出した値（ヒント用）。None の場合はヒントなし。
        client: google.genai.Client インスタンス。コンフリクト解決にLLMを使う場合に指定。

    Returns:
        {技術文書キー: {項目名: 統合後の値}}
    """
    field_names = [f["name"] for f in fields]
    catalog = catalog_data or {}

    # 技術文書タイプごとに全抽出データを収集
    doc_type_items: dict[str, list[dict]] = {}
    for img_name, result in results.items():
        data_list = result.get("data", [])
        if not data_list:
            continue

        cls_info = classification.get(img_name, {})
        category = cls_info.get("category", "")
        doc_type_key = DOC_TYPE_MAP.get(category, "")
        if not doc_type_key:
            continue

        doc_type_items.setdefault(doc_type_key, []).extend(data_list)

    # 各技術文書タイプで値を統合（コンフリクトを検出・収集）
    merged: dict[str, dict] = {}
    # コンフリクト情報を収集（LLMでまとめて解決するため）
    all_conflicts: list[dict] = []
    # コンフリクトの位置情報: (doc_type_key, field_name) → conflict_index
    conflict_positions: dict[tuple[str, str], int] = {}

    for doc_type_key, items_list in doc_type_items.items():
        combined: dict[str, str] = {}

        for name in field_names:
            # 全ページから有効な値を収集
            valid_values: list[str] = []
            has_unclear = False
            for item in items_list:
                val = item.get(name, "-") or "-"
                if val == "不鮮明":
                    has_unclear = True
                elif val != "-":
                    valid_values.append(val)

            if not valid_values:
                # 有効値なし
                combined[name] = "不鮮明" if has_unclear else "-"
            else:
                # 正規化後の重複を除去
                seen: dict[str, str] = {}  # normalized → original
                unique_values: list[str] = []
                for v in valid_values:
                    norm = normalize_for_comparison(v).lower()
                    if norm not in seen:
                        seen[norm] = v
                        unique_values.append(v)

                if len(unique_values) == 1:
                    # 値が1種類 → そのまま採用
                    combined[name] = unique_values[0]
                else:
                    # コンフリクト → まず収集、後でまとめてLLM判定
                    hint = catalog.get(name, "-")
                    conflict_idx = len(all_conflicts) + 1
                    all_conflicts.append({
                        "index": conflict_idx,
                        "field": name,
                        "doc_type": doc_type_key,
                        "candidates": unique_values,
                        "catalog_hint": hint,
                    })
                    conflict_positions[(doc_type_key, name)] = conflict_idx
                    # 仮にフォールバック値を設定（後でLLM結果で上書き）
                    combined[name] = _resolve_conflict(unique_values, hint)

        merged[doc_type_key] = combined

    # LLMによるコンフリクト解決
    if all_conflicts and client is not None:
        logger.info("マージコンフリクト %d 件をLLMで解決中...", len(all_conflicts))
        llm_selections = _resolve_conflict_with_llm(client, all_conflicts)

        # LLM結果を反映
        for conflict in all_conflicts:
            idx = conflict["index"]
            if idx in llm_selections:
                sel = llm_selections[idx]
                candidates = conflict["candidates"]
                if 0 <= sel < len(candidates):
                    doc_key = conflict["doc_type"]
                    field_name = conflict["field"]
                    old_val = merged[doc_key][field_name]
                    merged[doc_key][field_name] = candidates[sel]
                    if candidates[sel] != old_val:
                        logger.info(
                            "  LLM解決: [%s] %s → '%s'（フォールバック: '%s'）",
                            doc_key, field_name, candidates[sel], old_val,
                        )
    elif all_conflicts:
        logger.info("マージコンフリクト %d 件（LLMクライアントなし → フォールバック）", len(all_conflicts))

    return merged


def compare_by_doc_type(
    results: dict,
    classification: dict,
    correct_data: dict,
    fields: list[dict],
    catalog_data: dict | None = None,
    client=None,
) -> list[dict]:
    """分類結果に基づいて技術文書タイプ別に比較を行う。

    同じ文書タイプの複数画像（例: 取説3ページ）は1つに統合してから比較する。
    統合時のコンフリクトはLLM判定（clientがある場合）またはカタログヒントで解決する。

    Args:
        results: {画像名: {"data": [...]}} - 抽出結果
        classification: {画像名: {"category": "承認図", ...}} - 画像分類結果
        correct_data: {技術文書: {項目名: 正解値}} - マスターExcelの正解データ
        fields: 項目リスト
        catalog_data: カタログから抽出した値（コンフリクト解決のヒント用、省略可）
        client: google.genai.Client インスタンス（コンフリクト解決にLLMを使う場合）

    Returns:
        比較結果の行リスト。技術文書タイプごとに1行。
    """
    field_names = [f["name"] for f in fields]
    comparison_rows: list[dict] = []

    # 同じ文書タイプの結果を統合（LLM + カタログヒントでコンフリクト解決）
    merged = _merge_by_doc_type(results, classification, fields, catalog_data, client)

    # 正解データ側のキーを正規化（表記ゆれ吸収）
    normalized_correct: dict[str, dict] = {}
    for key, val in correct_data.items():
        normalized_key = _CORRECT_KEY_ALIASES.get(key, key)
        normalized_correct[normalized_key] = val

    for doc_type_key, extracted_combined in merged.items():
        # 正解データから対応する技術文書のデータを取得
        correct_entry = normalized_correct.get(doc_type_key, {})
        if not correct_entry:
            continue

        # 品番を取得
        hinban = extracted_combined.get(field_names[0], "-") if field_names else "-"

        items: dict = {}
        for name in field_names:
            extracted_val = extracted_combined.get(name, "-")
            correct_val = correct_entry.get(name, "") if correct_entry else ""

            status = determine_match_status(extracted_val, correct_val, name)
            items[name] = {
                "correct": correct_val,
                "extracted": extracted_val,
                "status": status,
                "diff_comment": "",
            }

        comparison_rows.append(
            {
                "品番": hinban,
                "技術文書": doc_type_key,
                "items": items,
            }
        )

    return comparison_rows



def _calc_accuracy_from_rows(
    comparison_rows: list[dict], field_names: list[str]
) -> tuple[dict, dict[str, dict], dict[str, dict]]:
    """comparison_rows (compare_by_doc_typeの返り値) から正答率を算出する。

    Returns:
        (total_stats, by_field_stats, by_doc_stats)
        各 stats は {"match": N, "partial": N, "mismatch": N, "na": N}
    """
    total = {"match": 0, "partial": 0, "mismatch": 0, "na": 0}
    by_field: dict[str, dict] = {
        name: {"match": 0, "partial": 0, "mismatch": 0, "na": 0}
        for name in field_names
    }
    by_doc: dict[str, dict] = {}

    for row_data in comparison_rows:
        doc_type = row_data.get("技術文書", "")
        if doc_type not in by_doc:
            by_doc[doc_type] = {"match": 0, "partial": 0, "mismatch": 0, "na": 0}

        items = row_data.get("items", {})
        for name in field_names:
            info = items.get(name, {})
            status = info.get("status", "na")
            total[status] += 1
            by_field[name][status] += 1
            by_doc[doc_type][status] += 1

    return total, by_field, by_doc


def _write_accuracy_sheet(
    wb: openpyxl.Workbook,
    comparison_rows: list[dict],
    field_names: list[str],
    sheet_title: str = "正答率サマリー",
) -> None:
    """Excelワークブックに正答率サマリーシートを追加する。"""
    total, by_field, by_doc = _calc_accuracy_from_rows(comparison_rows, field_names)

    ws = wb.create_sheet(title=sheet_title[:31])

    # スタイル定義
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    fill_header = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    fill_ok = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    fill_ng = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    fill_partial = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    pct_fmt = '0.0%'
    center = Alignment(horizontal="center")

    def _accuracy(stats: dict) -> float:
        """正答率を返す。N/A除外。"""
        cnt = stats["match"] + stats["partial"] + stats["mismatch"]
        if cnt == 0:
            return 0.0
        return stats["match"] / cnt

    row = 1
    # --- タイトル ---
    ws.cell(row=row, column=1, value="正答率サマリー").font = title_font
    row += 2

    # --- 【全体】 ---
    ws.cell(row=row, column=1, value="【全体】").font = section_font
    row += 1

    overall_headers = ["OK(完全一致)", "部分一致", "NG(不一致)", "N/A", "比較項目数(N/A除外)", "正答率"]
    for ci, h in enumerate(overall_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
    row += 1

    cnt_excl_na = total["match"] + total["partial"] + total["mismatch"]
    accuracy = _accuracy(total)
    values = [total["match"], total["partial"], total["mismatch"], total["na"], cnt_excl_na, accuracy]
    for ci, v in enumerate(values, 1):
        c = ws.cell(row=row, column=ci, value=v)
        c.alignment = center
        if ci == 6:
            c.number_format = pct_fmt
    # 正答率セルに色付け
    rate_cell = ws.cell(row=row, column=6)
    if accuracy >= 0.9:
        rate_cell.fill = fill_ok
    elif accuracy >= 0.7:
        rate_cell.fill = fill_partial
    else:
        rate_cell.fill = fill_ng
    row += 2

    # --- 【項目別】 ---
    ws.cell(row=row, column=1, value="【項目別正答率】").font = section_font
    row += 1

    field_headers = ["項目", "OK", "部分一致", "NG", "N/A", "比較数", "正答率"]
    for ci, h in enumerate(field_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
    row += 1

    for name in field_names:
        stats = by_field.get(name, {"match": 0, "partial": 0, "mismatch": 0, "na": 0})
        cnt = stats["match"] + stats["partial"] + stats["mismatch"]
        acc = _accuracy(stats)
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=stats["match"]).alignment = center
        ws.cell(row=row, column=3, value=stats["partial"]).alignment = center
        ws.cell(row=row, column=4, value=stats["mismatch"]).alignment = center
        ws.cell(row=row, column=5, value=stats["na"]).alignment = center
        ws.cell(row=row, column=6, value=cnt).alignment = center
        c7 = ws.cell(row=row, column=7, value=acc)
        c7.number_format = pct_fmt
        c7.alignment = center
        if cnt > 0:
            if acc >= 0.9:
                c7.fill = fill_ok
            elif acc >= 0.7:
                c7.fill = fill_partial
            else:
                c7.fill = fill_ng
        row += 1
    row += 1

    # --- 【技術文書別】 ---
    ws.cell(row=row, column=1, value="【技術文書別正答率】").font = section_font
    row += 1

    doc_headers = ["技術文書", "OK", "部分一致", "NG", "N/A", "比較数", "正答率"]
    for ci, h in enumerate(doc_headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
    row += 1

    for doc_type, stats in by_doc.items():
        cnt = stats["match"] + stats["partial"] + stats["mismatch"]
        acc = _accuracy(stats)
        ws.cell(row=row, column=1, value=doc_type)
        ws.cell(row=row, column=2, value=stats["match"]).alignment = center
        ws.cell(row=row, column=3, value=stats["partial"]).alignment = center
        ws.cell(row=row, column=4, value=stats["mismatch"]).alignment = center
        ws.cell(row=row, column=5, value=stats["na"]).alignment = center
        ws.cell(row=row, column=6, value=cnt).alignment = center
        c7 = ws.cell(row=row, column=7, value=acc)
        c7.number_format = pct_fmt
        c7.alignment = center
        if cnt > 0:
            if acc >= 0.9:
                c7.fill = fill_ok
            elif acc >= 0.7:
                c7.fill = fill_partial
            else:
                c7.fill = fill_ng
        row += 1

    # 列幅調整
    for ci in range(1, 8):
        col_letter = openpyxl.utils.get_column_letter(ci)
        max_len = 0
        for r in ws.iter_rows(min_col=ci, max_col=ci, min_row=1, max_row=ws.max_row):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len * 1.5 + 2, 10)

def export_extraction_excel(
    results: dict, classification: dict, fields: list[dict], output_path: str
) -> None:
    """抽出結果をExcelファイルに出力する。

    Args:
        results: {画像名: {"data": [...]}} - 抽出結果
        classification: {画像名: {"category": ...}} - 画像分類結果
        fields: 項目リスト
        output_path: 出力ファイルパス
    """
    field_names = [f["name"] for f in fields]
    fill_header = PatternFill(
        start_color="D5D8DC", end_color="D5D8DC", fill_type="solid"
    )
    header_font = Font(bold=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "読み取り結果"

    headers = ["画像名", "技術文書"] + field_names
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    for img_name, result in results.items():
        data_list = result.get("data", [])
        cls_info = classification.get(img_name, {})
        category = cls_info.get("category", "")

        if not data_list:
            ws.cell(row=row_idx, column=1, value=img_name)
            ws.cell(row=row_idx, column=2, value=category)
            row_idx += 1
            continue

        for item in data_list:
            ws.cell(row=row_idx, column=1, value=img_name)
            ws.cell(row=row_idx, column=2, value=category)
            for fi, name in enumerate(field_names):
                ws.cell(row=row_idx, column=3 + fi, value=item.get(name, "-") or "-")
            row_idx += 1

    # 列幅調整
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
        ):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        ws.column_dimensions[col_letter].width = max(max_length * 1.5 + 2, 8)

    wb.save(output_path)
    wb.close()


# 技術文書の表示順序と表示名
_DOC_TYPE_ORDER = ["図面", "承認図", "器具銘板", "取扱説明書", "外装ラベル"]
_DOC_TYPE_DISPLAY = {
    "図面": "図面(組立図)",
    "承認図": "承認図",
    "器具銘板": "器具銘板(現物)",
    "取扱説明書": "取説",
    "外装ラベル": "外装表示",
}


def _auto_adjust_column_widths(ws) -> None:
    """シートの列幅を内容に合わせて自動調整する。"""
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in ws.iter_rows(
            min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row
        ):
            for cell in row:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
        ws.column_dimensions[col_letter].width = max(max_length * 1.5 + 2, 8)


def _write_vertical_comparison_sheet(
    ws,
    comparison_rows: list[dict],
    field_names: list[str],
    catalog_data: dict | None = None,
) -> None:
    """縦並びレイアウトで比較結果シートを書き込む。

    行=項目、列=技術文書タイプ。
    カタログは「抽出」の1列のみ。
    他の技術文書は「抽出」「正解」「判定(○/×)」の3列セット。
    差分コメントがある場合はセルコメントとして付与。

    ヘッダー構成（2行）:
      行1: 項目名 | カタログ | ──図面(組立図)── | ──承認図── | ...
      行2:        |         | 抽出 | 正解 | 判定  | 抽出 | 正解 | 判定 | ...
    """
    catalog = catalog_data or {}

    # スタイル定義
    fill_ok = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    fill_ng = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    fill_header = PatternFill(
        start_color="D5D8DC", end_color="D5D8DC", fill_type="solid"
    )
    fill_catalog = PatternFill(
        start_color="DAEEF3", end_color="DAEEF3", fill_type="solid"
    )
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_fill = {"match": fill_ok, "mismatch": fill_ng}
    status_symbol = {"match": "○", "mismatch": "×", "na": "-"}

    # comparison_rows を技術文書タイプ別に整理
    by_doc: dict[str, dict] = {}
    for row_data in comparison_rows:
        doc_type = row_data.get("技術文書", "")
        by_doc[doc_type] = row_data.get("items", {})

    # 存在する技術文書タイプを順序通りに取得
    doc_types = [dt for dt in _DOC_TYPE_ORDER if dt in by_doc]

    # --- ヘッダー行1: 項目名 | カタログ | 技術文書名（3列結合） | ... ---
    col = 1
    cell = ws.cell(row=1, column=col, value="項目名")
    cell.font = header_font
    cell.fill = fill_header
    cell.alignment = center
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    col = 2
    cell = ws.cell(row=1, column=col, value="カタログ")
    cell.font = header_font
    cell.fill = fill_catalog
    cell.alignment = center
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    doc_col_start: dict[str, int] = {}  # 技術文書 → 開始列
    col = 3
    for dt in doc_types:
        display_name = _DOC_TYPE_DISPLAY.get(dt, dt)
        doc_col_start[dt] = col
        cell = ws.cell(row=1, column=col, value=display_name)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = center
        # 3列を結合
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)

        # ヘッダー行2: 抽出 | 正解 | 判定
        for sub_idx, sub_label in enumerate(["抽出", "正解", "判定"]):
            sub_cell = ws.cell(row=2, column=col + sub_idx, value=sub_label)
            sub_cell.font = header_font
            sub_cell.fill = fill_header
            sub_cell.alignment = center

        col += 3

    # --- データ行: 各項目を1行ずつ（3行目から） ---
    data_start_row = 3
    for row_idx, name in enumerate(field_names, data_start_row):
        # 項目名
        ws.cell(row=row_idx, column=1, value=name)

        # カタログ値
        cat_val = catalog.get(name, "-") or "-"
        cat_cell = ws.cell(row=row_idx, column=2, value=cat_val)
        if cat_val and cat_val != "-":
            cat_cell.fill = fill_catalog

        # 各技術文書: 抽出 | 正解 | 判定
        for dt in doc_types:
            base_col = doc_col_start[dt]
            items = by_doc.get(dt, {})
            info = items.get(name, {})
            extracted = info.get("extracted", "-") or "-"
            correct = info.get("correct", "-") or "-"
            status = info.get("status", "na")
            diff_comment = info.get("diff_comment", "")

            # 抽出値
            ext_cell = ws.cell(row=row_idx, column=base_col, value=extracted)
            ext_cell.alignment = Alignment(vertical="center", wrap_text=True)
            if status in status_fill:
                ext_cell.fill = status_fill[status]
            if diff_comment:
                ext_cell.comment = Comment(diff_comment, "自動生成")

            # 正解値
            ws.cell(row=row_idx, column=base_col + 1, value=correct)

            # 判定（○/×/-）
            judge_cell = ws.cell(
                row=row_idx, column=base_col + 2, value=status_symbol.get(status, "-")
            )
            judge_cell.alignment = center
            if status in status_fill:
                judge_cell.fill = status_fill[status]

    _auto_adjust_column_widths(ws)



def export_comparison_excel(
    comparison_rows: list[dict],
    fields: list[dict],
    output_path: str,
    catalog_data: dict | None = None,
) -> bytes:
    """比較結果をExcelファイルに出力する（縦並びレイアウト）。

    行=項目、列=技術文書タイプ。カタログ列を一番左に配置。

    Args:
        comparison_rows: compare_by_doc_typeの返り値
        fields: 項目リスト
        output_path: 出力ファイルパス
        catalog_data: カタログから抽出した値（省略可）

    Returns:
        Excelファイルのバイトデータ（st.download_button用）
    """
    field_names = [f["name"] for f in fields]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "比較結果"

    _write_vertical_comparison_sheet(ws, comparison_rows, field_names, catalog_data)

    # 正答率サマリーシートを追加
    _write_accuracy_sheet(wb, comparison_rows, field_names)

    # ファイルに保存
    wb.save(output_path)

    # BytesIOにも書き込んでbytesとして返す
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()


def export_multi_product_comparison_excel(
    all_comparisons: dict[str, list[dict]],
    all_fields: dict[str, list[dict]],
    output_path: str,
    all_catalog_data: dict[str, dict] | None = None,
) -> bytes:
    """全品番の正誤結果を1つのExcelにまとめて出力する（縦並びレイアウト）。

    品番ごとにシートを分けて出力する。各シートは縦並びレイアウト
    （行=項目、列=技術文書タイプ）で、カタログ列を一番左に配置。

    Args:
        all_comparisons: {品番: compare_by_doc_typeの返り値} のdict
        all_fields: {品番: 項目リスト} のdict。品番ごとに異なる項目に対応
        output_path: 出力ファイルパス
        all_catalog_data: {品番: カタログ抽出値} のdict（省略可）

    Returns:
        Excelファイルのバイトデータ（st.download_button用）
    """
    catalogs = all_catalog_data or {}

    wb = openpyxl.Workbook()
    first_sheet = True

    for product_id, comparison_rows in all_comparisons.items():
        product_fields = all_fields.get(product_id, [])
        field_names = [f["name"] for f in product_fields]
        catalog_data = catalogs.get(product_id)

        # シート名はExcelの制限で31文字まで
        sheet_name = product_id[:31] if len(product_id) > 31 else product_id

        if first_sheet:
            ws = wb.active
            ws.title = sheet_name
            first_sheet = False
        else:
            ws = wb.create_sheet(title=sheet_name)

        _write_vertical_comparison_sheet(ws, comparison_rows, field_names, catalog_data)

    # 全体サマリーシートを追加
    all_rows = []
    merged_field_names = []
    for product_id, comparison_rows in all_comparisons.items():
        product_fields = all_fields.get(product_id, [])
        for fn in [f["name"] for f in product_fields]:
            if fn not in merged_field_names:
                merged_field_names.append(fn)
        all_rows.extend(comparison_rows)

    # 品番別正答率シート
    ws_summary = wb.create_sheet(title="全体サマリー")
    title_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    fill_header = PatternFill(start_color="D5D8DC", end_color="D5D8DC", fill_type="solid")
    fill_ok = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    fill_ng = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    fill_partial = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    pct_fmt = "0.0%"
    center = Alignment(horizontal="center")

    row = 1
    ws_summary.cell(row=row, column=1, value="全品番 正答率サマリー").font = title_font
    row += 2

    ws_summary.cell(row=row, column=1, value="【品番別正答率】").font = section_font
    row += 1

    prod_headers = ["品番", "OK", "部分一致", "NG", "N/A", "比較数", "正答率"]
    for ci, h in enumerate(prod_headers, 1):
        c = ws_summary.cell(row=row, column=ci, value=h)
        c.font = header_font
        c.fill = fill_header
        c.alignment = center
    row += 1

    grand_total = {"match": 0, "partial": 0, "mismatch": 0, "na": 0}

    for product_id, comparison_rows in all_comparisons.items():
        product_fields = all_fields.get(product_id, [])
        p_field_names = [f["name"] for f in product_fields]
        total, _, _ = _calc_accuracy_from_rows(comparison_rows, p_field_names)

        for k in grand_total:
            grand_total[k] += total[k]

        cnt = total["match"] + total["partial"] + total["mismatch"]
        acc = total["match"] / cnt if cnt > 0 else 0.0

        ws_summary.cell(row=row, column=1, value=product_id)
        ws_summary.cell(row=row, column=2, value=total["match"]).alignment = center
        ws_summary.cell(row=row, column=3, value=total["partial"]).alignment = center
        ws_summary.cell(row=row, column=4, value=total["mismatch"]).alignment = center
        ws_summary.cell(row=row, column=5, value=total["na"]).alignment = center
        ws_summary.cell(row=row, column=6, value=cnt).alignment = center
        c7 = ws_summary.cell(row=row, column=7, value=acc)
        c7.number_format = pct_fmt
        c7.alignment = center
        if cnt > 0:
            if acc >= 0.9:
                c7.fill = fill_ok
            elif acc >= 0.7:
                c7.fill = fill_partial
            else:
                c7.fill = fill_ng
        row += 1

    # 全体合計行
    row += 1
    g_cnt = grand_total["match"] + grand_total["partial"] + grand_total["mismatch"]
    g_acc = grand_total["match"] / g_cnt if g_cnt > 0 else 0.0

    ws_summary.cell(row=row, column=1, value="全体合計").font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=grand_total["match"]).alignment = center
    ws_summary.cell(row=row, column=3, value=grand_total["partial"]).alignment = center
    ws_summary.cell(row=row, column=4, value=grand_total["mismatch"]).alignment = center
    ws_summary.cell(row=row, column=5, value=grand_total["na"]).alignment = center
    ws_summary.cell(row=row, column=6, value=g_cnt).alignment = center
    c7 = ws_summary.cell(row=row, column=7, value=g_acc)
    c7.number_format = pct_fmt
    c7.alignment = center
    c7.font = Font(bold=True)
    if g_cnt > 0:
        if g_acc >= 0.9:
            c7.fill = fill_ok
        elif g_acc >= 0.7:
            c7.fill = fill_partial
        else:
            c7.fill = fill_ng

    # 列幅調整
    for ci in range(1, 8):
        col_letter = openpyxl.utils.get_column_letter(ci)
        max_len = 0
        for r in ws_summary.iter_rows(min_col=ci, max_col=ci, min_row=1, max_row=ws_summary.max_row):
            for cell in r:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
        ws_summary.column_dimensions[col_letter].width = max(max_len * 1.5 + 2, 10)


    # ファイルに保存
    wb.save(output_path)

    # BytesIOにも書き込んでbytesとして返す
    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer.getvalue()
